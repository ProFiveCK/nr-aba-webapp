"""
reformat_payroll.py
-------------------
Reformats the TechOne / FMIS SuperPaymentFile payroll deduction export into a
clean summary workbook.

Sheets produced:
  - One sheet per Pay Period End Date (pure data rows, import-safe)
  - "Summary" sheet with per-period and grand totals

Columns per data sheet:
  ID Number | Family Name | Given Name | Title | Date of Birth |
  Pay Period End Date | Employee Contribution | Employer Contribution
"""

import argparse
import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE  = os.path.join(SCRIPT_DIR, "SuperPaymentFile export_20260630122329.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "Payroll Deduction Summary - June 2026.xlsx")

# ── style constants ───────────────────────────────────────────────────────────
HDR_FILL     = PatternFill("solid", fgColor="1F4E79")
HDR_FONT     = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TOTAL_FILL   = PatternFill("solid", fgColor="D6E4F0")
TOTAL_FONT   = Font(name="Calibri", bold=True, size=11)
GRAND_FILL   = PatternFill("solid", fgColor="1F4E79")
GRAND_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CURRENCY_FMT = "#,##0.00"
CENTER       = Alignment(horizontal="center")
RIGHT        = Alignment(horizontal="right")

COL_WIDTHS = {
    "ID Number": 12, "Family Name": 25, "Given Name": 20,
    "Title": 8, "Date of Birth": 14,
    "Pay Period End Date": 20,
    "Employee Contribution": 22, "Employer Contribution": 22,
}

# ── helper: style a data sheet ───────────────────────────────────────────────
def style_data_sheet(ws, df_period):
    # Header
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER

    # Column widths
    for idx, col_name in enumerate(df_period.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS.get(col_name, 15)

    # Data rows — date format on cols 5 (Date of Birth) & 6 (Pay Period End Date),
    # currency format on cols 7 & 8
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in (5, 6) and cell.value not in (None, ""):
                cell.number_format = "DD-MM-YYYY"
            if cell.column in (7, 8):
                cell.alignment = RIGHT
                if cell.value is not None:
                    cell.number_format = CURRENCY_FMT

    # Total row
    total_row = ws.max_row + 1
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL

    label = ws.cell(row=total_row, column=6, value="TOTAL")
    label.font = TOTAL_FONT
    label.fill = TOTAL_FILL
    label.alignment = RIGHT

    last_data = total_row - 1
    for col in (7, 8):
        col_letter = get_column_letter(col)
        tc = ws.cell(row=total_row, column=col,
                     value=f"=SUM({col_letter}2:{col_letter}{last_data})")
        tc.font = TOTAL_FONT
        tc.fill = TOTAL_FILL
        tc.alignment = RIGHT
        tc.number_format = CURRENCY_FMT


def derive_output_path(input_path):
    input_dir = os.path.dirname(os.path.abspath(input_path))
    input_name = os.path.basename(input_path)
    if input_name.lower().endswith(".xlsx"):
        stem = input_name[:-5]
    else:
        stem = os.path.splitext(input_name)[0]
    return os.path.join(input_dir, f"{stem} - Reformatted.xlsx")


def reformat_payroll(input_path, output_path, sheet_name="Sheet1"):
    # ── 1. load & transform ───────────────────────────────────────────────────
    # FMIS SuperPaymentFile export has 5 preamble rows (title, parameters,
    # criteria, blank, then the column headers) so the header is on row 6 (header=5).
    # If the requested sheet does not exist, fall back to the first available sheet.
    xl = pd.ExcelFile(input_path)
    if sheet_name not in xl.sheet_names:
        sheet_name = xl.sheet_names[0]

    df = pd.read_excel(input_path, sheet_name=sheet_name, header=5)
    df.columns = df.columns.str.strip()

    df["Pay Component - Code"] = pd.to_numeric(df["Pay Component - Code"], errors="coerce")
    df = df[df["Pay Component - Code"].isin([825.0, 1000.0])].copy()

    df["Pay Period End Date"] = pd.to_datetime(df["Pay Period End Date"]).dt.date
    df["Date of Birth"] = pd.to_datetime(df["Date of Birth"], errors="coerce").dt.date
    df["Title"] = df["Title"].fillna("")
    df["Date of Birth"] = df["Date of Birth"].where(df["Date of Birth"].notna(), "")

    code_map = {825.0: "Employee Contribution", 1000.0: "Employer Contribution"}
    df["Pay Component - Code"] = df["Pay Component - Code"].map(code_map)

    group_cols = ["Employee Id", "Last Name", "First Name", "Title",
                  "Date of Birth", "Pay Period End Date"]

    pivoted = df.pivot_table(
        index=group_cols,
        columns="Pay Component - Code",
        values="Amount",
        aggfunc="sum"
    ).reset_index()
    pivoted.columns.name = None

    for col in ["Employee Contribution", "Employer Contribution"]:
        if col not in pivoted.columns:
            pivoted[col] = None

    final = pivoted[["Employee Id", "Last Name", "First Name", "Title",
                     "Date of Birth",
                     "Pay Period End Date", "Employee Contribution", "Employer Contribution"]]
    final = final.rename(columns={"Employee Id": "ID Number",
                                  "Last Name": "Family Name",
                                  "First Name": "Given Name"})
    final = final.sort_values(["Pay Period End Date", "ID Number"]).reset_index(drop=True)

    periods = sorted(final["Pay Period End Date"].unique())

    # ── 2. write workbook ─────────────────────────────────────────────────────
    summary_data = []

    with pd.ExcelWriter(output_path, engine="openpyxl", date_format="DD-MM-YYYY") as writer:
        for period in periods:
            df_period = final[final["Pay Period End Date"] == period].copy()
            period_sheet_name = str(period)

            df_period.to_excel(writer, sheet_name=period_sheet_name, index=False)
            ws = writer.sheets[period_sheet_name]
            style_data_sheet(ws, df_period)

            summary_data.append({
                "Pay Period End Date": period,
                "Employees": len(df_period),
                "Employee Contribution": df_period["Employee Contribution"].sum(),
                "Employer Contribution": df_period["Employer Contribution"].sum(),
            })

        df_summary = pd.DataFrame(summary_data)
        df_summary["Total Contribution"] = (
            df_summary["Employee Contribution"] + df_summary["Employer Contribution"]
        )

        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        ws_sum = writer.sheets["Summary"]

        for cell in ws_sum[1]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER

        sum_widths = [20, 12, 22, 22, 20]
        for idx, width in enumerate(sum_widths, start=1):
            ws_sum.column_dimensions[get_column_letter(idx)].width = width

        for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row):
            for cell in row:
                if cell.column == 1 and cell.value is not None:
                    cell.number_format = "DD-MM-YYYY"
                if cell.column in (3, 4, 5):
                    cell.alignment = RIGHT
                    if cell.value is not None:
                        cell.number_format = CURRENCY_FMT

        grand_row = ws_sum.max_row + 1
        last_data = grand_row - 1

        for col in range(1, 6):
            ws_sum.cell(row=grand_row, column=col).fill = GRAND_FILL

        label = ws_sum.cell(row=grand_row, column=1, value="GRAND TOTAL")
        label.font = GRAND_FONT
        label.fill = GRAND_FILL

        emp_count = ws_sum.cell(row=grand_row, column=2, value=f"=SUM(B2:B{last_data})")
        emp_count.font = GRAND_FONT
        emp_count.fill = GRAND_FILL
        emp_count.alignment = RIGHT

        for col in (3, 4, 5):
            col_letter = get_column_letter(col)
            gc = ws_sum.cell(row=grand_row, column=col, value=f"=SUM({col_letter}2:{col_letter}{last_data})")
            gc.font = GRAND_FONT
            gc.fill = GRAND_FILL
            gc.alignment = RIGHT
            gc.number_format = CURRENCY_FMT

        wb = writer.book
        wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)
        wb.active = wb["Summary"]

    return periods, summary_data


def parse_args():
    parser = argparse.ArgumentParser(description="Reformat a TechOne/FMIS payroll deduction workbook.")
    parser.add_argument("--input", default=INPUT_FILE, help="Path to the source Excel workbook.")
    parser.add_argument("--output", default=None, help="Path to write the reformatted workbook.")
    parser.add_argument("--sheet", default="Sheet1", help="Source sheet name to read.")
    return parser.parse_args()


def main():
    args = parse_args()
    input_path = os.path.abspath(args.input)
    output_path = os.path.abspath(args.output) if args.output else derive_output_path(input_path)
    periods, summary_data = reformat_payroll(input_path, output_path, args.sheet)

    print(f"Done — {len(periods)} pay period sheet(s) + Summary sheet written to:\n  {output_path}")
    for row in summary_data:
        print(f"  {row['Pay Period End Date']}  |  {row['Employees']:>4} employees  "
              f"|  Emp: {row['Employee Contribution']:>10,.2f}  "
              f"|  Employer: {row['Employer Contribution']:>10,.2f}")


if __name__ == "__main__":
    main()
